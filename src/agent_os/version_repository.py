"""
PMAIOS v0.3.5 多产品线版本库管理器

核心功能:
- 用 Excel 管理多产品线版本迭代记录
- 自动记录各产品线的版本变更
- 支持查看版本历史、对比差异
- 保存每次版本更新记录
"""
import json
import os
from datetime import datetime
from typing import Dict, Any, List, Optional
from dataclasses import dataclass, asdict
from enum import Enum
import hashlib


# ============================================================
# 数据模型
# ============================================================

class VersionType(str, Enum):
    """版本类型"""
    MAJOR = "major"      # 主版本号 x.0.0
    MINOR = "minor"      # 次版本号 0.x.0
    PATCH = "patch"      # 修订号 0.0.x
    HOTFIX = "hotfix"    # 热修复


class ReleaseStatus(str, Enum):
    """发布状态"""
    DRAFT = "draft"              # 草稿
    IN_PROGRESS = "in_progress"  # 开发中
    TESTING = "testing"          # 测试中
    RELEASED = "released"        # 已发布
    DEPRECATED = "deprecated"    # 已废弃


@dataclass
class VersionRecord:
    """版本记录"""
    id: str
    product_line: str
    version: str           # 如 "v0.3.5"
    version_type: str      # major/minor/patch/hotfix
    release_status: str    # draft/in_progress/testing/released
    release_date: str      # 发布日期
    created_at: str = ""
    updated_at: str = ""
    summary: str = ""      # 版本摘要
    changes: str = ""      # JSON 数组，变更列表
    breaking_changes: str = ""  # 破坏性变更说明
    migration_needed: bool = False  # 是否需要迁移
    artifacts: str = ""    # 相关产物路径，逗号分隔
    commit_hash: str = ""  # Git commit
    tags: str = ""         # 标签，逗号分隔

    def __post_init__(self):
        now = datetime.utcnow().isoformat()
        if not self.created_at:
            self.created_at = now
        if not self.updated_at:
            self.updated_at = now
        if not self.release_date:
            self.release_date = datetime.utcnow().strftime("%Y-%m-%d")

    def to_excel_row(self) -> Dict[str, Any]:
        """转换为 Excel 行数据"""
        return {
            "ID": self.id,
            "产品线": self.product_line,
            "版本号": self.version,
            "版本类型": self.version_type,
            "发布状态": self.release_status,
            "发布日期": self.release_date,
            "创建时间": self.created_at,
            "更新时间": self.updated_at,
            "版本摘要": self.summary,
            "变更内容": self.changes,
            "破坏性变更": self.breaking_changes,
            "需要迁移": self.migration_needed,
            "相关产物": self.artifacts,
            "Commit Hash": self.commit_hash,
            "标签": self.tags,
        }

    @classmethod
    def from_excel_row(cls, row: Dict[str, Any]) -> "VersionRecord":
        """从 Excel 行创建"""
        return cls(
            id=row.get("ID", ""),
            product_line=row.get("产品线", ""),
            version=row.get("版本号", ""),
            version_type=row.get("版本类型", "patch"),
            release_status=row.get("发布状态", "draft"),
            release_date=row.get("发布日期", ""),
            created_at=row.get("创建时间", ""),
            updated_at=row.get("更新时间", ""),
            summary=row.get("版本摘要", ""),
            changes=row.get("变更内容", ""),
            breaking_changes=row.get("破坏性变更", ""),
            migration_needed=bool(row.get("需要迁移", False)),
            artifacts=row.get("相关产物", ""),
            commit_hash=row.get("Commit Hash", ""),
            tags=row.get("标签", ""),
        )


@dataclass
class ChangeLog:
    """变更日志"""
    id: str
    version_id: str
    product_line: str
    timestamp: str
    change_type: str     # feature/fix/docs/refactor/test
    module: str          # 模块名称
    description: str     # 变更描述
    author: str          # 作者
    files_changed: str   # 变更文件，分号分隔
    lines_added: int = 0
    lines_deleted: int = 0

    def to_excel_row(self) -> Dict[str, Any]:
        return {
            "ID": self.id,
            "版本 ID": self.version_id,
            "产品线": self.product_line,
            "时间": self.timestamp,
            "变更类型": self.change_type,
            "模块": self.module,
            "描述": self.description,
            "作者": self.author,
            "变更文件": self.files_changed,
            "新增行数": self.lines_added,
            "删除行数": self.lines_deleted,
        }


# ============================================================
# Excel 管理器
# ============================================================

class VersionRepository:
    """多产品线版本库管理器"""

    def __init__(self, excel_path: str = "./version_repository.xlsx"):
        self.excel_path = excel_path
        self.records_path = excel_path.replace(".xlsx", "_records.xlsx")

        # 延迟导入 openpyxl
        try:
            from openpyxl import Workbook, load_workbook
            self.openpyxl = __import__('openpyxl')
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        self._ensure_excel_exists()

    def _ensure_excel_exists(self):
        """确保 Excel 文件存在"""
        if not os.path.exists(self.excel_path):
            self._create_excel()

    def _create_excel(self):
        """创建 Excel 文件"""
        wb = self.openpyxl.Workbook()

        # Sheet1: 版本记录总表
        ws_versions = wb.active
        ws_versions.title = "版本记录"
        headers = ["ID", "产品线", "版本号", "版本类型", "发布状态", "发布日期",
                   "创建时间", "更新时间", "版本摘要", "变更内容", "破坏性变更",
                   "需要迁移", "相关产物", "Commit Hash", "标签"]
        ws_versions.append(headers)

        # Sheet2: 变更日志
        ws_changelog = wb.create_sheet("变更日志")
        changelog_headers = ["ID", "版本 ID", "产品线", "时间", "变更类型", "模块",
                            "描述", "作者", "变更文件", "新增行数", "删除行数"]
        ws_changelog.append(changelog_headers)

        # Sheet3: 产品线概览
        ws_summary = wb.create_sheet("产品线概览")
        summary_headers = ["产品线", "最新版本", "版本总数", "已发布版本",
                          "最后更新时间", "维护者", "状态"]
        ws_summary.append(summary_headers)

        wb.save(self.excel_path)
        print(f"Created version repository: {self.excel_path}")

    def add_version(self, record: VersionRecord) -> str:
        """添加版本记录"""
        wb = self.openpyxl.load_workbook(self.excel_path)
        ws = wb["版本记录"]

        row = record.to_excel_row()
        ws.append([
            row["ID"], row["产品线"], row["版本号"], row["版本类型"],
            row["发布状态"], row["发布日期"], row["创建时间"], row["更新时间"],
            row["版本摘要"], row["变更内容"], row["破坏性变更"],
            row["需要迁移"], row["相关产物"], row["Commit Hash"], row["标签"]
        ])

        wb.save(self.excel_path)
        return record.id

    def add_change_log(self, changelog: ChangeLog) -> str:
        """添加变更日志"""
        wb = self.openpyxl.load_workbook(self.excel_path)
        ws = wb["变更日志"]

        row = changelog.to_excel_row()
        ws.append([
            row["ID"], row["版本 ID"], row["产品线"], row["时间"],
            row["变更类型"], row["模块"], row["描述"], row["作者"],
            row["变更文件"], row["新增行数"], row["删除行数"]
        ])

        wb.save(self.excel_path)
        return changelog.id

    def get_versions_by_product(self, product_line: str) -> List[VersionRecord]:
        """获取指定产品线的所有版本"""
        wb = self.openpyxl.load_workbook(self.excel_path)
        ws = wb["版本记录"]

        versions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == product_line:  # 产品线列
                versions.append(VersionRecord(
                    id=str(row[0]),
                    product_line=str(row[1]),
                    version=str(row[2]),
                    version_type=str(row[3]),
                    release_status=str(row[4]),
                    release_date=str(row[5]) if row[5] else "",
                    created_at=str(row[6]) if row[6] else "",
                    updated_at=str(row[7]) if row[7] else "",
                    summary=str(row[8]) if row[8] else "",
                    changes=str(row[9]) if row[9] else "",
                    breaking_changes=str(row[10]) if row[10] else "",
                    migration_needed=bool(row[11]),
                    artifacts=str(row[12]) if row[12] else "",
                    commit_hash=str(row[13]) if row[13] else "",
                    tags=str(row[14]) if row[14] else ""
                ))

        return sorted(versions, key=lambda x: x.version, reverse=True)

    def get_latest_version(self, product_line: str) -> Optional[VersionRecord]:
        """获取指定产品线的最新版本"""
        versions = self.get_versions_by_product(product_line)
        return versions[0] if versions else None

    def get_all_product_lines(self) -> List[str]:
        """获取所有产品线"""
        wb = self.openpyxl.load_workbook(self.excel_path)
        ws = wb["版本记录"]

        product_lines = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                product_lines.add(str(row[1]))

        return sorted(list(product_lines))

    def get_version_history(self, product_line: str, version: str) -> List[ChangeLog]:
        """获取指定版本的变更历史"""
        wb = self.openpyxl.load_workbook(self.excel_path)
        ws = wb["变更日志"]

        change_logs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == product_line and row[1] == version:
                change_logs.append(ChangeLog(
                    id=str(row[0]),
                    version_id=str(row[1]),
                    product_line=str(row[2]),
                    timestamp=str(row[3]) if row[3] else "",
                    change_type=str(row[4]),
                    module=str(row[5]),
                    description=str(row[6]),
                    author=str(row[7]),
                    files_changed=str(row[8]) if row[8] else "",
                    lines_added=int(row[9]) if row[9] else 0,
                    lines_deleted=int(row[10]) if row[10] else 0
                ))

        return change_logs

    def get_summary(self) -> Dict[str, Any]:
        """获取版本库摘要"""
        wb = self.openpyxl.load_workbook(self.excel_path)
        ws = wb["版本记录"]

        product_stats = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            pl = str(row[1]) if row[1] else "unknown"
            if pl not in product_stats:
                product_stats[pl] = {
                    "total": 0,
                    "released": 0,
                    "latest_version": "",
                    "last_updated": ""
                }

            product_stats[pl]["total"] += 1
            if row[4] == "released":
                product_stats[pl]["released"] += 1

            version = str(row[2]) if row[2] else ""
            if version > product_stats[pl]["latest_version"]:
                product_stats[pl]["latest_version"] = version

            updated = str(row[7]) if row[7] else ""
            if updated > product_stats[pl]["last_updated"]:
                product_stats[pl]["last_updated"] = updated

        return {
            "product_lines": list(product_stats.keys()),
            "stats": product_stats,
            "total_versions": sum(s["total"] for s in product_stats.values()),
        }

    def update_summary_sheet(self):
        """更新产品线概览表"""
        wb = self.openpyxl.load_workbook(self.excel_path)
        ws = wb["产品线概览"]

        # 清空现有数据（保留表头）
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        summary = self.get_summary()
        for pl, stats in summary["stats"].items():
            ws.append([
                pl,
                stats["latest_version"],
                stats["total"],
                stats["released"],
                stats["last_updated"],
                "",  # 维护者
                "active"  # 状态
            ])

        wb.save(self.excel_path)

    def auto_record_from_git(self, product_line: str, commit_hash: str,
                             changes: List[Dict[str, Any]]) -> str:
        """从 Git commit 自动记录版本变更"""
        version_id = f"{product_line.replace('/', '-')}-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

        # 创建版本记录（如果不存在则更新）
        latest = self.get_latest_version(product_line)
        if latest:
            # 更新现有版本
            record = latest
            record.updated_at = datetime.utcnow().isoformat()
            record.changes = json.dumps(changes, ensure_ascii=False)
            record.commit_hash = commit_hash
        else:
            # 创建新版本
            record = VersionRecord(
                id=version_id,
                product_line=product_line,
                version="v0.1.0",
                version_type="minor",
                release_status="in_progress",
                release_date="",
                summary=f"Auto-recorded from commit {commit_hash[:8]}",
                changes=json.dumps(changes, ensure_ascii=False),
                commit_hash=commit_hash
            )

        self.add_version(record)

        # 添加变更日志
        for change in changes:
            changelog = ChangeLog(
                id=f"{version_id}-{hashlib.md5(change.get('description', '').encode()).hexdigest()[:8]}",
                version_id=version_id,
                product_line=product_line,
                timestamp=datetime.utcnow().isoformat(),
                change_type=change.get("type", "feature"),
                module=change.get("module", "unknown"),
                description=change.get("description", ""),
                author=change.get("author", "auto"),
                files_changed=";".join(change.get("files", [])),
                lines_added=change.get("lines_added", 0),
                lines_deleted=change.get("lines_deleted", 0)
            )
            self.add_change_log(changelog)

        self.update_summary_sheet()
        return version_id


# ============================================================
# 便捷函数
# ============================================================

def create_version(product_line: str, version: str, summary: str,
                   changes: List[str], release_status: str = "draft") -> VersionRecord:
    """创建版本记录"""
    return VersionRecord(
        id=f"{product_line.replace('/', '-')}-{version.replace('.', '-')}",
        product_line=product_line,
        version=version,
        version_type="minor",
        release_status=release_status,
        release_date="",
        summary=summary,
        changes=json.dumps(changes, ensure_ascii=False),
        tags="auto-generated"
    )


def create_change_log(version_id: str, product_line: str,
                      change_type: str, module: str,
                      description: str, files: List[str]) -> ChangeLog:
    """创建变更日志"""
    return ChangeLog(
        id=f"{version_id}-{hashlib.md5(description.encode()).hexdigest()[:8]}",
        version_id=version_id,
        product_line=product_line,
        timestamp=datetime.utcnow().isoformat(),
        change_type=change_type,
        module=module,
        description=description,
        author="auto",
        files_changed=";".join(files)
    )


if __name__ == "__main__":
    # 测试示例
    repo = VersionRepository("./version_repository.xlsx")

    # 获取摘要
    summary = repo.get_summary()
    print(f"版本库摘要：{json.dumps(summary, indent=2, ensure_ascii=False)}")

    # 获取所有产品线
    product_lines = repo.get_all_product_lines()
    print(f"产品线列表：{product_lines}")
