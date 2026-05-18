"""
PMAIOS v0.3.4 Excel 需求池管理器

核心功能:
- 用 Excel 管理多产品线需求池
- 从会议纪要、日报、OKR 自动提取需求并写入 Excel
- 根据项目文档自动收敛
- 保存每次更新记录到 Excel
"""
import json
import os
from datetime import datetime
from typing import Dict, Any, List, Optional, Set
from dataclasses import dataclass, asdict, field
from enum import Enum
import hashlib

ANTHROPIC_MODEL = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")


def create_anthropic_client():
    import anthropic
    return anthropic.Anthropic(api_key=ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else anthropic.Anthropic()


# ============================================================
# 数据模型
# ============================================================

class RequirementSource(str, Enum):
    """需求来源"""
    MEETING = "meeting"
    DAILY_REPORT = "daily_report"
    OKR = "okr"
    PROJECT_DOC = "project_doc"
    MANUAL = "manual"


class RequirementStatus(str, Enum):
    """需求状态"""
    DRAFT = "draft"
    BACKLOG = "backlog"
    ANALYZING = "analyzing"
    READY = "ready"
    IN_PROGRESS = "in_progress"
    DONE = "done"
    ARCHIVED = "archived"


class RequirementPriority(str, Enum):
    """需求优先级"""
    P0 = "P0"
    P1 = "P1"
    P2 = "P2"
    P3 = "P3"


@dataclass
class Requirement:
    """需求定义"""
    id: str
    title: str
    description: str
    product_line: str
    source: str  # RequirementSource 的值
    source_doc_id: str
    status: str = "draft"  # RequirementStatus 的值
    priority: str = "P2"  # RequirementPriority 的值
    created_at: str = ""
    updated_at: str = ""
    tags: str = ""  # Excel 中用逗号分隔
    related_requirements: str = ""  # Excel 中用逗号分隔
    owner: str = ""
    acceptance_criteria: str = ""  # Excel 中用分号分隔
    confidence: float = 0.5

    def __post_init__(self):
        now = datetime.utcnow().isoformat()
        if not self.created_at:
            self.created_at = now
        if not self.updated_at:
            self.updated_at = now

    def to_excel_row(self) -> Dict[str, Any]:
        """转换为 Excel 行数据"""
        return {
            "ID": self.id,
            "标题": self.title,
            "描述": self.description,
            "产品线": self.product_line,
            "来源": self.source,
            "来源文档 ID": self.source_doc_id,
            "状态": self.status,
            "优先级": self.priority,
            "创建时间": self.created_at,
            "更新时间": self.updated_at,
            "标签": self.tags,
            "关联需求": self.related_requirements,
            "负责人": self.owner,
            "验收标准": self.acceptance_criteria,
            "置信度": self.confidence,
        }

    @classmethod
    def from_excel_row(cls, row: Dict[str, Any]) -> "Requirement":
        """从 Excel 行创建"""
        return cls(
            id=row.get("ID", ""),
            title=row.get("标题", ""),
            description=row.get("描述", ""),
            product_line=row.get("产品线", ""),
            source=row.get("来源", "manual"),
            source_doc_id=row.get("来源文档 ID", ""),
            status=row.get("状态", "draft"),
            priority=row.get("优先级", "P2"),
            created_at=row.get("创建时间", ""),
            updated_at=row.get("更新时间", ""),
            tags=row.get("标签", ""),
            related_requirements=row.get("关联需求", ""),
            owner=row.get("负责人", ""),
            acceptance_criteria=row.get("验收标准", ""),
            confidence=float(row.get("置信度", 0.5)),
        )


@dataclass
class UpdateRecord:
    """更新记录"""
    id: str
    timestamp: str
    update_type: str
    requirements_affected: str  # Excel 中用逗号分隔
    source_doc_id: str
    changes: str  # Excel 中 JSON 字符串
    summary: str

    def to_excel_row(self) -> Dict[str, Any]:
        return {
            "ID": self.id,
            "时间": self.timestamp,
            "类型": self.update_type,
            "影响需求": self.requirements_affected,
            "来源文档": self.source_doc_id,
            "变更详情": self.changes,
            "摘要": self.summary,
        }


# ============================================================
# Excel 管理器
# ============================================================

class ExcelRequirementPool:
    """Excel 需求池管理器"""

    def __init__(self, excel_path: str = "./requirement_pool.xlsx"):
        self.excel_path = excel_path
        self.records_excel_path = excel_path.replace(".xlsx", "_records.xlsx")

        # 初始化 Excel 文件
        self._init_excel()

    def _init_excel(self):
        """初始化 Excel 文件"""
        try:
            from openpyxl import Workbook
        except ImportError:
            print("请安装 openpyxl: pip install openpyxl")
            raise

        # 主需求池
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "需求池"

            # 表头
            headers = [
                "ID", "标题", "描述", "产品线", "来源", "来源文档 ID",
                "状态", "优先级", "创建时间", "更新时间",
                "标签", "关联需求", "负责人", "验收标准", "置信度"
            ]
            ws.append(headers)

            # 设置列宽
            column_widths = [15, 30, 50, 20, 15, 20, 12, 8, 20, 20, 20, 20, 15, 40, 10]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width

            # 数据验证 - 状态
            from openpyxl.worksheet.datavalidation import DataValidation
            status_dv = DataValidation(type="list", formula1='"draft,backlog,analyzing,ready,in_progress,done,archived"')
            ws.add_data_validation(status_dv)
            status_dv.add(f"G2:G1000")

            # 数据验证 - 优先级
            priority_dv = DataValidation(type="list", formula1='"P0,P1,P2,P3"')
            ws.add_data_validation(priority_dv)
            priority_dv.add(f"H2:H1000")

            wb.save(self.excel_path)
            print(f"创建需求池 Excel: {self.excel_path}")

        # 更新记录表
        if not os.path.exists(self.records_excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "更新记录"

            headers = ["ID", "时间", "类型", "影响需求", "来源文档", "变更详情", "摘要"]
            ws.append(headers)

            column_widths = [15, 20, 12, 40, 20, 60, 50]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width

            wb.save(self.records_excel_path)
            print(f"创建更新记录 Excel: {self.records_excel_path}")

    def _read_excel(self, sheet_name: str = None):
        """读取 Excel 数据"""
        from openpyxl import load_workbook
        wb = load_workbook(self.excel_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        # 转换为字典列表
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if row[i] is not None}
            rows.append(row_dict)
        return rows

    def _write_requirement(self, req: Requirement):
        """写入需求到 Excel"""
        from openpyxl import load_workbook
        wb = load_workbook(self.excel_path)
        ws = wb.active

        # 检查是否已存在
        existing_row = None
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row[0] == req.id:
                existing_row = i
                break

        row_data = list(req.to_excel_row().values())

        if existing_row:
            # 更新现有行
            for col, value in enumerate(row_data, 1):
                ws.cell(row=existing_row, column=col, value=value)
        else:
            # 添加新行
            ws.append(row_data)

        wb.save(self.excel_path)

    def _write_record(self, record: UpdateRecord):
        """写入更新记录到 Excel"""
        from openpyxl import load_workbook
        wb = load_workbook(self.records_excel_path)
        ws = wb.active
        ws.append(list(record.to_excel_row().values()))
        wb.save(self.records_excel_path)

    def _generate_id(self, prefix: str = "") -> str:
        """生成 ID"""
        timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
        return f"{prefix}{timestamp}" if prefix else timestamp

    def _parse_json_list(self, text: str) -> List[Dict[str, Any]]:
        """解析 JSON 列表"""
        try:
            start = text.find("[")
            end = text.rfind("]") + 1
            if start >= 0 and end > start:
                return json.loads(text[start:end])
        except:
            pass
        return []

    def _parse_json(self, text: str) -> Dict[str, Any]:
        """解析 JSON 对象"""
        try:
            start = text.find("{")
            end = text.rfind("}") + 1
            if start >= 0 and end > start:
                return json.loads(text[start:end])
        except:
            pass
        return {}

    # ============================================================
    # 核心 API
    # ============================================================

    def extract_from_meeting(self, meeting_notes: str, product_line: str, meeting_id: str = "") -> List[Requirement]:
        """从会议纪要提取需求"""
        client = create_anthropic_client()

        if not meeting_id:
            meeting_id = self._generate_id("mtg_")

        prompt = f"""从以下会议纪要中提取产品需求：

会议纪要：
{meeting_notes}

产品线：{product_line}

返回 JSON 列表，每个需求包含：title, description, priority(P0/P1/P2/P3), tags(数组), acceptance_criteria(数组), confidence(0-1)
"""
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )

        output = response.content[0].text if response.content else "[]"
        requirements_data = self._parse_json_list(output)

        created = []
        for req_data in requirements_data:
            req = Requirement(
                id=self._generate_id("req_"),
                title=req_data.get("title", ""),
                description=req_data.get("description", ""),
                product_line=product_line,
                source="meeting",
                source_doc_id=meeting_id,
                priority=req_data.get("priority", "P2"),
                tags=",".join(req_data.get("tags", [])),
                acceptance_criteria=";".join(req_data.get("acceptance_criteria", [])),
                confidence=req_data.get("confidence", 0.5),
            )
            self._write_requirement(req)
            created.append(req)

        if created:
            record = UpdateRecord(
                id=self._generate_id("rec_"),
                timestamp=datetime.utcnow().isoformat(),
                update_type="create",
                requirements_affected=",".join([r.id for r in created]),
                source_doc_id=meeting_id,
                changes=json.dumps({"count": len(created)}, ensure_ascii=False),
                summary=f"从会议纪要提取 {len(created)} 个需求 - 产品线：{product_line}"
            )
            self._write_record(record)

        return created

    def extract_from_daily_report(self, report_content: str, author: str, product_line: str, report_id: str = "") -> List[Requirement]:
        """从日报提取需求"""
        client = create_anthropic_client()

        if not report_id:
            report_id = self._generate_id("daily_")

        prompt = f"""从以下日报中提取潜在的产品需求：

日报内容：
{report_content}

作者：{author}
产品线：{product_line}

返回 JSON 列表，每个需求包含：title, description, priority(P1/P2/P3), tags(数组), confidence(0-1)
"""
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}]
        )

        output = response.content[0].text if response.content else "[]"
        requirements_data = self._parse_json_list(output)

        created = []
        for req_data in requirements_data:
            req = Requirement(
                id=self._generate_id("req_"),
                title=req_data.get("title", ""),
                description=req_data.get("description", ""),
                product_line=product_line,
                source="daily_report",
                source_doc_id=report_id,
                priority=req_data.get("priority", "P2"),
                tags=",".join(req_data.get("tags", [])),
                confidence=req_data.get("confidence", 0.5),
                owner=author,
            )
            self._write_requirement(req)
            created.append(req)

        if created:
            record = UpdateRecord(
                id=self._generate_id("rec_"),
                timestamp=datetime.utcnow().isoformat(),
                update_type="create",
                requirements_affected=",".join([r.id for r in created]),
                source_doc_id=report_id,
                changes=json.dumps({"count": len(created), "author": author}, ensure_ascii=False),
                summary=f"从 {author} 的日报提取 {len(created)} 个需求 - 产品线：{product_line}"
            )
            self._write_record(record)

        return created

    def extract_from_okr(self, okr_data: Dict[str, Any], product_line: str, okr_id: str = "") -> List[Requirement]:
        """从 OKR 提取需求"""
        client = create_anthropic_client()

        if not okr_id:
            okr_id = self._generate_id("okr_")

        prompt = f"""从以下 OKR 中提取产品需求：

OKR 数据：
{json.dumps(okr_data, ensure_ascii=False, indent=2)}

产品线：{product_line}

返回 JSON 列表，每个需求包含：title, description(关联到哪个 O 或 KR), priority(P0/P1/P2), tags(数组), acceptance_criteria(数组), confidence(0-1)
"""
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )

        output = response.content[0].text if response.content else "[]"
        requirements_data = self._parse_json_list(output)

        created = []
        for req_data in requirements_data:
            req = Requirement(
                id=self._generate_id("req_"),
                title=req_data.get("title", ""),
                description=req_data.get("description", ""),
                product_line=product_line,
                source="okr",
                source_doc_id=okr_id,
                priority=req_data.get("priority", "P1"),
                tags=",".join(req_data.get("tags", [])),
                acceptance_criteria=";".join(req_data.get("acceptance_criteria", [])),
                confidence=req_data.get("confidence", 0.7),
            )
            self._write_requirement(req)
            created.append(req)

        if created:
            record = UpdateRecord(
                id=self._generate_id("rec_"),
                timestamp=datetime.utcnow().isoformat(),
                update_type="create",
                requirements_affected=",".join([r.id for r in created]),
                source_doc_id=okr_id,
                changes=json.dumps({"count": len(created)}, ensure_ascii=False),
                summary=f"从 OKR 提取 {len(created)} 个需求 - 产品线：{product_line}"
            )
            self._write_record(record)

        return created

    def converge_requirements(self, project_docs: List[Dict[str, Any]]) -> List[Requirement]:
        """根据项目文档收敛需求"""
        client = create_anthropic_client()

        # 读取现有需求
        rows = self._read_excel()
        draft_requirements = [
            Requirement.from_excel_row(r)
            for r in rows
            if r.get("状态") in ["draft", "backlog"]
        ]

        if not draft_requirements:
            return []

        prompt = f"""分析以下需求和项目文档，进行需求收敛和合并：

现有需求池 ({len(draft_requirements)} 个):
{json.dumps([Requirement.from_excel_row(r).to_excel_row() for r in rows[:20]], ensure_ascii=False, indent=2)}

项目文档:
{json.dumps(project_docs, ensure_ascii=False, indent=2)}

请识别：
1. 重复或相似的需求 - 返回 merges 数组
2. 需要调整优先级的需求 - 返回 priority_updates 数组
3. 需要新增的需求 - 返回 new_requirements 数组

返回 JSON 格式：
{{
  "merges": [{{"target": "req_xxx", "sources": ["req_aaa"], "reason": "..."}}],
  "priority_updates": [{{"id": "req_xxx", "new_priority": "P0", "reason": "..."}}],
  "new_requirements": [{{"title": "...", "description": "...", "priority": "P1"}}]
}}
"""
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}]
        )

        output = response.content[0].text if response.content else "{}"
        result = self._parse_json(output)

        affected_ids = []

        # 执行合并
        for merge in result.get("merges", []):
            target_id = merge.get("target")
            sources = merge.get("sources", [])
            # 在 Excel 中标记合并关系
            for row in self._read_excel():
                if row.get("ID") == target_id:
                    existing_related = row.get("关联需求", "")
                    new_related = existing_related + "," + ",".join(sources) if existing_related else ",".join(sources)
                    # 更新 Excel
                    self._update_excel_cell("ID", target_id, "关联需求", new_related)
                    affected_ids.extend(sources + [target_id])
                    break

        # 更新优先级
        for update in result.get("priority_updates", []):
            req_id = update.get("id")
            new_priority = update.get("new_priority")
            if req_id and new_priority:
                self._update_excel_cell("ID", req_id, "优先级", new_priority)
                affected_ids.append(req_id)

        # 创建新需求
        for new_req in result.get("new_requirements", []):
            req = Requirement(
                id=self._generate_id("req_"),
                title=new_req.get("title", ""),
                description=new_req.get("description", ""),
                product_line="",
                source="project_doc",
                source_doc_id="auto_converge",
                priority=new_req.get("priority", "P2"),
                confidence=0.6
            )
            self._write_requirement(req)
            affected_ids.append(req.id)

        # 记录更新
        if affected_ids:
            record = UpdateRecord(
                id=self._generate_id("rec_"),
                timestamp=datetime.utcnow().isoformat(),
                update_type="converge",
                requirements_affected=",".join(affected_ids),
                source_doc_id="project_docs",
                changes=json.dumps(result, ensure_ascii=False),
                summary=f"需求收敛：影响 {len(affected_ids)} 个需求"
            )
            self._write_record(record)

        return [Requirement.from_excel_row(r) for r in self._read_excel() if r.get("ID") in affected_ids]

    def _update_excel_cell(self, lookup_column: str, lookup_value: Any, target_column: str, new_value: Any):
        """更新 Excel 单元格"""
        from openpyxl import load_workbook
        wb = load_workbook(self.excel_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        lookup_col_idx = headers.index(lookup_column) + 1
        target_col_idx = headers.index(target_column) + 1

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row[lookup_col_idx - 1] == lookup_value:
                ws.cell(row=row_num, column=target_col_idx, value=new_value)
                break

        wb.save(self.excel_path)

    # ============================================================
    # 查询 API
    # ============================================================

    def get_requirements(self, product_line: Optional[str] = None, status: Optional[str] = None) -> List[Requirement]:
        """查询需求"""
        rows = self._read_excel()
        results = [Requirement.from_excel_row(r) for r in rows]

        if product_line:
            results = [r for r in results if r.product_line == product_line]
        if status:
            results = [r for r in results if r.status == status]

        return results

    def get_summary(self) -> Dict[str, Any]:
        """获取摘要"""
        rows = self._read_excel()
        requirements = [Requirement.from_excel_row(r) for r in rows]

        product_lines = set(r.product_line for r in requirements if r.product_line)

        return {
            "total": len(requirements),
            "product_lines": list(product_lines),
            "by_status": {
                status: len([r for r in requirements if r.status == status])
                for status in ["draft", "backlog", "analyzing", "ready", "in_progress", "done", "archived"]
            },
            "by_priority": {
                priority: len([r for r in requirements if r.priority == priority])
                for priority in ["P0", "P1", "P2", "P3"]
            },
        }

    def export_to_excel(self, output_path: str):
        """导出到指定 Excel 文件"""
        from openpyxl import load_workbook
        wb = load_workbook(self.excel_path)
        wb.save(output_path)
        print(f"已导出到：{output_path}")


# ============================================================
# 使用示例
# ============================================================

if __name__ == "__main__":
    # 创建需求池
    pool = ExcelRequirementPool("./requirement_pool.xlsx")

    print("需求池已创建！")
    print(f"主文件：{pool.excel_path}")
    print(f"记录文件：{pool.records_excel_path}")

    # 示例：获取摘要
    summary = pool.get_summary()
    print(f"\n需求池摘要：{summary}")
