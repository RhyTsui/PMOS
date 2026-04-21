from __future__ import annotations

import csv
import io
import json
import zlib
from datetime import datetime
from pathlib import Path

import duckdb
from openpyxl import load_workbook
from unstructured.partition.csv import partition_csv
from unstructured.partition.xlsx import partition_xlsx


INBOX = Path(r"E:\AI\ai-os\docs\sources\inbox")
OUT_DIR = Path(r"E:\AI\ai-os\subprojects\tracking-acceptance\docs\sources\index")
DB_PATH = OUT_DIR / "inbox.duckdb"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").strip()


def stable_doc_id(path: Path) -> str:
    return f"{path.stem}_{int(path.stat().st_mtime)}_{zlib.crc32(str(path).encode('utf-8')):08x}"


def safe_console(text: str) -> str:
    return text.encode("gbk", errors="backslashreplace").decode("gbk")


def load_workbook_from_bytes(path: Path):
    data = path.read_bytes()
    return load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def bootstrap_schema(conn: duckdb.DuckDBPyConnection) -> None:
    conn.execute(
        """
        create table if not exists documents (
          doc_id varchar primary key,
          file_name varchar,
          file_path varchar,
          extension varchar,
          mtime timestamp,
          mtime_epoch double,
          size_bytes bigint
        )
        """
    )
    conn.execute("alter table documents add column if not exists mtime_epoch double")
    conn.execute(
        """
        create table if not exists sheets (
          doc_id varchar,
          sheet_name varchar,
          sheet_index integer,
          max_row integer,
          max_column integer
        )
        """
    )
    conn.execute(
        """
        create table if not exists row_chunks (
          doc_id varchar,
          source_type varchar,
          sheet_name varchar,
          row_number integer,
          column_count integer,
          row_text varchar,
          row_json varchar
        )
        """
    )
    conn.execute(
        """
        create table if not exists cell_chunks (
          doc_id varchar,
          source_type varchar,
          sheet_name varchar,
          row_number integer,
          column_number integer,
          column_label varchar,
          cell_text varchar
        )
        """
    )
    conn.execute(
        """
        create table if not exists elements (
          doc_id varchar,
          source_type varchar,
          sheet_name varchar,
          element_index integer,
          element_type varchar,
          text varchar,
          html varchar,
          metadata_json varchar
        )
        """
    )
    conn.execute("create index if not exists idx_documents_name on documents(file_name)")
    conn.execute("create index if not exists idx_rows_doc_sheet on row_chunks(doc_id, sheet_name)")
    conn.execute("create index if not exists idx_cells_doc_sheet on cell_chunks(doc_id, sheet_name)")


def replace_document(conn: duckdb.DuckDBPyConnection, path: Path) -> str:
    doc_id = stable_doc_id(path)
    conn.execute("delete from sheets where doc_id = ?", [doc_id])
    conn.execute("delete from row_chunks where doc_id = ?", [doc_id])
    conn.execute("delete from cell_chunks where doc_id = ?", [doc_id])
    conn.execute("delete from elements where doc_id = ?", [doc_id])
    conn.execute("delete from documents where doc_id = ?", [doc_id])
    conn.execute(
        """
        insert into documents(doc_id, file_name, file_path, extension, mtime, mtime_epoch, size_bytes)
        values (?, ?, ?, ?, ?, ?, ?)
        """,
        [
            doc_id,
            path.name,
            str(path),
            path.suffix.lower(),
            datetime.fromtimestamp(path.stat().st_mtime),
            path.stat().st_mtime,
            path.stat().st_size,
        ],
    )
    return doc_id


def ingest_xlsx(conn: duckdb.DuckDBPyConnection, path: Path) -> None:
    doc_id = replace_document(conn, path)
    wb = load_workbook_from_bytes(path)

    for sheet_index, ws_name in enumerate(wb.sheetnames, 1):
        ws = wb[ws_name]
        conn.execute(
            "insert into sheets values (?, ?, ?, ?, ?)",
            [doc_id, ws_name, sheet_index, ws.max_row, ws.max_column],
        )

        for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
            values = [normalize_text(v) for v in row]
            row_text = " | ".join(v for v in values if v)
            conn.execute(
                "insert into row_chunks values (?, ?, ?, ?, ?, ?, ?)",
                [doc_id, "xlsx", ws_name, row_number, len(values), row_text, json.dumps(values, ensure_ascii=False)],
            )
            for column_number, cell_text in enumerate(values, 1):
                if not cell_text:
                    continue
                conn.execute(
                    "insert into cell_chunks values (?, ?, ?, ?, ?, ?, ?)",
                    [doc_id, "xlsx", ws_name, row_number, column_number, f"C{column_number}", cell_text],
                )

    for element_index, el in enumerate(partition_xlsx(filename=str(path)), 1):
        metadata = getattr(el, "metadata", None)
        sheet_name = getattr(metadata, "page_name", None)
        html = getattr(metadata, "text_as_html", None)
        metadata_json = {}
        if metadata is not None and hasattr(metadata, "to_dict"):
            metadata_json = metadata.to_dict()
        conn.execute(
            "insert into elements values (?, ?, ?, ?, ?, ?, ?, ?)",
            [
                doc_id,
                "xlsx",
                sheet_name,
                element_index,
                type(el).__name__,
                normalize_text(getattr(el, "text", "")),
                html,
                json.dumps(metadata_json, ensure_ascii=False),
            ],
        )


def ingest_csv(conn: duckdb.DuckDBPyConnection, path: Path) -> None:
    doc_id = replace_document(conn, path)
    text = path.read_text(encoding="utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    sheet_name = "__csv__"
    max_cols = max((len(row) for row in rows), default=0)
    conn.execute(
        "insert into sheets values (?, ?, ?, ?, ?)",
        [doc_id, sheet_name, 1, len(rows), max_cols],
    )

    for row_number, row in enumerate(rows, 1):
        values = [normalize_text(v) for v in row]
        row_text = " | ".join(v for v in values if v)
        conn.execute(
            "insert into row_chunks values (?, ?, ?, ?, ?, ?, ?)",
            [doc_id, "csv", sheet_name, row_number, len(values), row_text, json.dumps(values, ensure_ascii=False)],
        )
        for column_number, cell_text in enumerate(values, 1):
            if not cell_text:
                continue
            conn.execute(
                "insert into cell_chunks values (?, ?, ?, ?, ?, ?, ?)",
                [doc_id, "csv", sheet_name, row_number, column_number, f"C{column_number}", cell_text],
            )

    for element_index, el in enumerate(partition_csv(filename=str(path)), 1):
        metadata = getattr(el, "metadata", None)
        html = getattr(metadata, "text_as_html", None)
        metadata_json = {}
        if metadata is not None and hasattr(metadata, "to_dict"):
            metadata_json = metadata.to_dict()
        conn.execute(
            "insert into elements values (?, ?, ?, ?, ?, ?, ?, ?)",
            [
                doc_id,
                "csv",
                sheet_name,
                element_index,
                type(el).__name__,
                normalize_text(getattr(el, "text", "")),
                html,
                json.dumps(metadata_json, ensure_ascii=False),
            ],
        )


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    conn = duckdb.connect(str(DB_PATH))
    bootstrap_schema(conn)

    files = sorted(
        [p for p in INBOX.iterdir() if p.suffix.lower() in {".xlsx", ".csv"}],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for path in files:
        if path.suffix.lower() == ".xlsx":
            ingest_xlsx(conn, path)
        elif path.suffix.lower() == ".csv":
            ingest_csv(conn, path)
        print(f"ingested\t{safe_console(path.name)}")

    summary = conn.execute(
        """
        select
          count(*) as documents,
          (select count(*) from sheets) as sheets,
          (select count(*) from row_chunks) as rows,
          (select count(*) from cell_chunks) as cells,
          (select count(*) from elements) as elements
        from documents
        """
    ).fetchone()
    print(
        json.dumps(
            {
                "db_path": str(DB_PATH),
                "documents": summary[0],
                "sheets": summary[1],
                "rows": summary[2],
                "cells": summary[3],
                "elements": summary[4],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
