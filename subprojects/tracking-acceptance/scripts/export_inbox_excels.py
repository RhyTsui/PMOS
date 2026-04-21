from __future__ import annotations

import csv
import io
import zlib
from pathlib import Path

from openpyxl import load_workbook


INBOX = Path(r"E:\AI\ai-os\docs\sources\inbox")
OUT_ROOT = Path(r"E:\AI\ai-os\subprojects\tracking-acceptance\docs\sources\converted")


def ascii_slug(text: str) -> str:
    slug = "".join(ch if ch.isascii() and ch.isalnum() else "_" for ch in text).strip("_")
    if slug:
        return slug
    return f"sheet_{zlib.crc32(text.encode('utf-8')):08x}"


def load_workbook_from_bytes(path: Path):
    data = path.read_bytes()
    return load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", " ").replace("\n", " ").strip()


def export_book(path: Path) -> Path:
    out_dir = OUT_ROOT / ascii_slug(path.stem)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook_from_bytes(path)

    summary_lines = [
        "# Excel 可读导出",
        "",
        f"源文件：`{path.name}`",
        "",
        f"工作表：{', '.join(wb.sheetnames)}",
        "",
    ]

    for idx, ws_name in enumerate(wb.sheetnames, 1):
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        safe = f"{idx:02d}_{ascii_slug(ws_name)}"
        csv_path = out_dir / f"{safe}.csv"
        md_path = out_dir / f"{safe}.md"

        with csv_path.open("w", encoding="utf-8-sig", newline="") as fp:
            writer = csv.writer(fp)
            for row in rows:
                writer.writerow([normalize_cell(v) for v in row])

        preview = rows[:80]
        max_cols = max((len(r) for r in preview), default=0)
        preview = [list(r) + [""] * (max_cols - len(r)) for r in preview]

        with md_path.open("w", encoding="utf-8") as fp:
            fp.write(f"# {ws_name}\n\n")
            fp.write(f"- 行数：{ws.max_row}\n")
            fp.write(f"- 列数：{ws.max_column}\n\n")
            if preview:
                header = [normalize_cell(x) for x in preview[0]]
                fp.write("|" + "|".join(header) + "|\n")
                fp.write("|" + "|".join(["---"] * max_cols) + "|\n")
                for row in preview[1:]:
                    vals = [normalize_cell(v) for v in row]
                    fp.write("|" + "|".join(vals) + "|\n")
            else:
                fp.write("_空表_\n")
            fp.write("\n> 注：Markdown 仅保留前 80 行预览；完整数据见同名 CSV。\n")

        summary_lines.append(
            f"- `{ws_name}`: {ws.max_row} 行 x {ws.max_column} 列 -> `{csv_path.name}` / `{md_path.name}`"
        )

    (out_dir / "README.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
    return out_dir


def main() -> None:
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    files = sorted(INBOX.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for path in files[:8]:
        out_dir = export_book(path)
        print(out_dir)


if __name__ == "__main__":
    main()
